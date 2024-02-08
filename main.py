import tkinter.messagebox
import openpyxl as px
from openpyxl.chart import Reference
from openpyxl.chart.axis import DateAxis
import glob
import os
import tkinter as tk
import tkinter.messagebox as messagebox
import sys

#tk.Tk().withdraw
#save_wb = 'result.xlsx'
#book = px.load_workbook(sum_wb)
#ws = book.active

save_wb = 'result.xlsx'
files = glob.glob('*.xlsx')
if save_wb in files:
    overwrite_check = tk.messagebox.askquestion(title='overwrite check', message='result.xlsxが存在します\n上書きしますか？')
    if overwrite_check == 'no':
        tk.messagebox.showwarning('abort', '処理を中止しました')
        exit()
else:
    create_file = px.Workbook()
    create_file.save(save_wb)

#集計用result.xlsxとシートを取得
book = px.load_workbook(save_wb)
sheet = book.active

#(工数)を含むファイル名のxlsxファイルを集計対象とする
input_wb = '*(工数)*.xlsx'

#必要そうなものをとりあえず書いとく
uketuke_sum = []
chousa_sum = []
anken_sum = []
settei_sum = []

#各ブックからセルの値をリストに格納
for i in glob.glob(input_wb):
    ws2 = px.load_workbook(i).worksheets[0]
    #受付のセルを集計 以下forは適当でbreakで調節する あとで書き直す
    for uketuke in range(2, 100, 1):
        cell_value = ws2.cell(row=uketuke, column=2).value
        if not cell_value == None : uketuke_sum.append(cell_value)
        #print(uketuke_sum) #debug
        if uketuke == 10: break

    #調査のセルを集計
    for chousa in range(11, 100, 1):
        cell_value = ws2.cell(row=chousa, column=2).value
        if not cell_value == None : chousa_sum.append(cell_value)
        #print(anken_sum) #debug
        if chousa == 19: break

    #案件のセルを集計
    for anken in range(20, 100, 1):
        cell_value = ws2.cell(row=anken, column=2).value
        if not cell_value == None : anken_sum.append(cell_value)
        #print(anken_sum) #debug
        if anken == 28: break

    #設定のセルを集計
    for settei in range(29, 100, 1):
        cell_value = ws2.cell(row=settei, column=2).value
        if not cell_value == None : settei_sum.append(cell_value)
        #print(anken_sum) #debug
        if settei == 38: break


result = sum(uketuke_sum)
result2 = sum(chousa_sum)
result3 = sum(anken_sum)
result4 = sum(settei_sum)
#print(result) #debug
#print(result2) #debug
#print(result3) #debug
#print(result4) #debug

#集計excelへ書き込み
sheet.cell(row=1, column=1).value = result
sheet.cell(row=2, column=1).value = result2
sheet.cell(row=3, column=1).value = result3
sheet.cell(row=4, column=1).value = result4

'''
#グラフ作成
graph_obj = px.chart.bar_chart()
graph_obj.title = '集計'
graph_obj.style = 12
graph_obj.height = 10
graph_obj.width = 15

graph_obj.y_axis.title = '工数'
#graph_obj.x_axis.number_format =
graph_obj.x_axis.title = '種別'

#Y軸範囲
y_axis = reference(sheet, min_col=2, min_row=3, max_col=2, max_row=9)
graph_obj.add.data(y_axis, title_from_data=True)

#X軸範囲
x_axis = reference(sheet, min_col=1, min_row=4, max_col=1, max_row=9)
graph_obj.set_categories(x_axis)

sheet.add_chart(graph_obj, 'E3')
'''

book.save(save_wb)
messagebox.showinfo('実行完了', '処理が完了しました。\nresult.xlsxが出力されていることを確認してください')
